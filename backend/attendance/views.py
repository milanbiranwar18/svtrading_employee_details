import base64
import uuid
import os
from decimal import Decimal
from datetime import date, timedelta
from calendar import monthrange

from django.core.files.base import ContentFile
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Sum, Count, Q

from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from employees.models import Employee
from .models import AttendanceRecord, LeaveRequest
from .serializers import AttendanceRecordSerializer, LeaveRequestSerializer


def save_base64_image(base64_string, folder):
    """Decode a base64 image string and save it to the media directory."""
    if not base64_string:
        return None
    try:
        if 'base64,' in base64_string:
            base64_string = base64_string.split('base64,')[1]
        img_data = base64.b64decode(base64_string)
        filename = f"{uuid.uuid4().hex}.jpg"
        path = os.path.join(folder, filename)
        return ContentFile(img_data, name=filename)
    except Exception:
        return None


class SignInView(APIView):
    """Public endpoint for employee sign-in."""
    permission_classes = [AllowAny]

    def post(self, request):
        employee_id = request.data.get('employee_id')
        photo_base64 = request.data.get('photo')
        location = request.data.get('location', '')

        if not employee_id:
            return Response({'error': 'employee_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = Employee.objects.get(id=employee_id, is_active=True)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found.'}, status=status.HTTP_404_NOT_FOUND)

        today = timezone.now().date()
        record, created = AttendanceRecord.objects.get_or_create(
            employee=employee, date=today,
            defaults={'sign_in_time': timezone.now(), 'sign_in_location': location, 'status': 'Present'}
        )

        if not created and record.sign_in_time:
            return Response({'error': 'Already signed in today.'}, status=status.HTTP_400_BAD_REQUEST)

        if not created:
            record.sign_in_time = timezone.now()
            record.sign_in_location = location
            record.status = 'Present'

        if photo_base64:
            photo_file = save_base64_image(photo_base64, 'attendance_photos/in/')
            if photo_file:
                record.sign_in_photo.save(photo_file.name, photo_file, save=False)

        record.save()
        serializer = AttendanceRecordSerializer(record)
        return Response(serializer.data, status=status.HTTP_200_OK)


class SignOutView(APIView):
    """Public endpoint for employee sign-out."""
    permission_classes = [AllowAny]

    def post(self, request):
        employee_id = request.data.get('employee_id')
        photo_base64 = request.data.get('photo')
        location = request.data.get('location', '')

        if not employee_id:
            return Response({'error': 'employee_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = Employee.objects.get(id=employee_id, is_active=True)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found.'}, status=status.HTTP_404_NOT_FOUND)

        today = timezone.now().date()
        try:
            record = AttendanceRecord.objects.get(employee=employee, date=today)
        except AttendanceRecord.DoesNotExist:
            return Response({'error': 'No sign-in record found for today.'}, status=status.HTTP_400_BAD_REQUEST)

        if not record.sign_in_time:
            return Response({'error': 'Employee has not signed in yet.'}, status=status.HTTP_400_BAD_REQUEST)

        if record.sign_out_time:
            return Response({'error': 'Already signed out today.'}, status=status.HTTP_400_BAD_REQUEST)

        record.sign_out_time = timezone.now()
        record.sign_out_location = location

        # Calculate total hours worked
        delta = record.sign_out_time - record.sign_in_time
        total_hours = Decimal(str(round(delta.total_seconds() / 3600, 2)))
        record.total_hours = total_hours

        # Determine status
        if total_hours < Decimal('4.0'):
            record.status = 'Absent'
        elif total_hours < Decimal('8.0'):
            record.status = 'Half Day'
        else:
            record.status = 'Present'

        if photo_base64:
            photo_file = save_base64_image(photo_base64, 'attendance_photos/out/')
            if photo_file:
                record.sign_out_photo.save(photo_file.name, photo_file, save=False)

        record.save()
        serializer = AttendanceRecordSerializer(record)
        return Response(serializer.data, status=status.HTTP_200_OK)


class TodayStatusView(APIView):
    """Public: Get today's attendance status for a specific employee."""
    permission_classes = [AllowAny]

    def get(self, request, employee_id):
        try:
            employee = Employee.objects.get(id=employee_id, is_active=True)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found.'}, status=status.HTTP_404_NOT_FOUND)

        today = timezone.now().date()
        try:
            record = AttendanceRecord.objects.get(employee=employee, date=today)
            serializer = AttendanceRecordSerializer(record)
            return Response(serializer.data)
        except AttendanceRecord.DoesNotExist:
            return Response({'message': 'No record for today.'}, status=status.HTTP_200_OK)


class MonthlyAttendanceView(APIView):
    """Public: Get current month attendance report for an employee."""
    permission_classes = [AllowAny]

    def get(self, request, employee_id):
        try:
            employee = Employee.objects.get(id=employee_id, is_active=True)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found.'}, status=status.HTTP_404_NOT_FOUND)

        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))
        _, num_days = monthrange(year, month)

        records = AttendanceRecord.objects.filter(
            employee=employee,
            date__year=year,
            date__month=month
        ).order_by('date')

        records_by_date = {r.date.day: r for r in records}

        result = []
        for day in range(1, num_days + 1):
            d = date(year, month, day)
            record = records_by_date.get(day)
            if record:
                serializer = AttendanceRecordSerializer(record)
                result.append(serializer.data)
            else:
                result.append({
                    'date': str(d),
                    'employee': employee.id,
                    'employee_name': employee.name,
                    'sign_in_time': None,
                    'sign_out_time': None,
                    'total_hours': None,
                    'status': 'Absent',
                })

        return Response({
            'employee': {'id': employee.id, 'name': employee.name, 'employee_code': employee.employee_code},
            'year': year,
            'month': month,
            'records': result
        })


class DashboardView(APIView):
    """Admin: Monthly summary of all employees."""
    permission_classes = [AllowAny]

    def get(self, request):
        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))

        employees = Employee.objects.filter(is_active=True).order_by('name')
        result = []

        for emp in employees:
            records = AttendanceRecord.objects.filter(
                employee=emp, date__year=year, date__month=month
            )
            present = records.filter(status='Present').count()
            half_day = records.filter(status='Half Day').count()
            leave = records.filter(status='Leave').count()
            total_hours = records.aggregate(total=Sum('total_hours'))['total'] or 0

            result.append({
                'id': emp.id,
                'name': emp.name,
                'employee_code': emp.employee_code,
                'present_days': present,
                'half_days': half_day,
                'leave_days': leave,
                'total_hours': float(total_hours),
            })

        return Response({'year': year, 'month': month, 'employees': result})


class ManualAttendanceView(APIView):
    """Admin: Create or update attendance record manually."""
    permission_classes = [AllowAny]

    def post(self, request):
        employee_id = request.data.get('employee_id')
        date_str = request.data.get('date')
        sign_in_time = request.data.get('sign_in_time')
        sign_out_time = request.data.get('sign_out_time')
        status_val = request.data.get('status', 'Present')
        location = request.data.get('location', 'Manual Entry')

        if not employee_id or not date_str:
            return Response({'error': 'employee_id and date are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            return Response({'error': 'Employee not found.'}, status=status.HTTP_404_NOT_FOUND)

        from datetime import datetime
        from django.utils.timezone import make_aware

        record, created = AttendanceRecord.objects.get_or_create(employee=employee, date=date_str)
        record.status = status_val
        record.sign_in_location = location
        record.sign_out_location = location

        if sign_in_time:
            try:
                dt = datetime.fromisoformat(f"{date_str}T{sign_in_time}")
                record.sign_in_time = make_aware(dt)
            except ValueError:
                pass

        if sign_out_time:
            try:
                dt = datetime.fromisoformat(f"{date_str}T{sign_out_time}")
                record.sign_out_time = make_aware(dt)
            except ValueError:
                pass

        if record.sign_in_time and record.sign_out_time:
            delta = record.sign_out_time - record.sign_in_time
            record.total_hours = Decimal(str(round(delta.total_seconds() / 3600, 2)))

        record.save()
        serializer = AttendanceRecordSerializer(record)
        return Response(serializer.data, status=status.HTTP_200_OK)


class AllAttendanceView(APIView):
    """Admin: Paginated attendance log for filtering."""
    permission_classes = [AllowAny]

    def get(self, request):
        queryset = AttendanceRecord.objects.all().order_by('-date')
        employee_id = request.query_params.get('employee_id')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)

        serializer = AttendanceRecordSerializer(queryset, many=True)
        return Response(serializer.data)


class LeaveRequestListCreateView(APIView):
    """Admin: List all leaves and create new ones."""
    permission_classes = [AllowAny]

    def get(self, request):
        leaves = LeaveRequest.objects.all().order_by('-created_at')
        serializer = LeaveRequestSerializer(leaves, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = LeaveRequestSerializer(data=request.data)
        if serializer.is_valid():
            leave = serializer.save()
            # Mark attendance records as Leave
            delta = leave.end_date - leave.start_date
            for i in range(delta.days + 1):
                d = leave.start_date + timedelta(days=i)
                record, _ = AttendanceRecord.objects.get_or_create(employee=leave.employee, date=d)
                record.status = 'Leave'
                record.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class LeaveRequestUpdateView(APIView):
    """Admin: Update leave request status (Approve/Reject)."""
    permission_classes = [AllowAny]

    def patch(self, request, pk):
        try:
            leave = LeaveRequest.objects.get(pk=pk)
        except LeaveRequest.DoesNotExist:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        leave.status = request.data.get('status', leave.status)
        leave.save()
        return Response(LeaveRequestSerializer(leave).data)


class ExportAttendanceView(APIView):
    """Admin: Export attendance data to Excel."""
    permission_classes = [AllowAny]

    def get(self, request):
        token_key = request.query_params.get('token')
        # If token is explicitly provided (and not string "null"), verify it.
        # Otherwise, since permission_classes = [AllowAny], allow the download
        # because the public /monthly-all page is unauthenticated.
        if token_key and token_key != 'null':
            from rest_framework.authtoken.models import Token
            try:
                token = Token.objects.get(key=token_key)
            except Token.DoesNotExist:
                return Response({"detail": "Invalid token."}, status=401)

        employee_id = request.query_params.get('employee_id')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))

        import calendar
        from datetime import date, timedelta
        
        # 1. Determine Date Range
        if date_from and date_to:
            start_date = date.fromisoformat(date_from)
            end_date = date.fromisoformat(date_to)
        elif date_from:
            start_date = date.fromisoformat(date_from)
            end_date = start_date
        else:
            _, last_day = calendar.monthrange(year, month)
            start_date = date(year, month, 1)
            end_date = date(year, month, last_day)

        num_days = (end_date - start_date).days + 1
        date_range = [start_date + timedelta(days=i) for i in range(num_days)]

        # 2. Fetch Employees and Attendance
        from employees.models import Employee
        employees = Employee.objects.filter(is_active=True).order_by('name')
        if employee_id:
            employees = employees.filter(id=employee_id)

        records = AttendanceRecord.objects.filter(date__gte=start_date, date__lte=end_date)
        attendance_dict = {(r.employee_id, r.date): r for r in records}

        # 3. Create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Matrix"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        
        # Static Headers + Dynamic Day Headers + Summary Headers
        headers = ['Employee Name', 'Code'] + [d.strftime('%d-%b') for d in date_range] + ['Present Days', 'Half Days', 'Leave Days', 'Absent Days', 'Total Hours']

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Set explicit column widths
            if col_num == 1:
                ws.column_dimensions[get_column_letter(col_num)].width = 20
            elif col_num == 2:
                ws.column_dimensions[get_column_letter(col_num)].width = 12
            elif col_num > 2 and col_num <= len(date_range) + 2:
                ws.column_dimensions[get_column_letter(col_num)].width = 18
            else:
                ws.column_dimensions[get_column_letter(col_num)].width = 14

        # Date formatter helper
        def fmt_time(dt):
            if dt:
                import pytz
                local_tz = pytz.timezone('Asia/Kolkata')
                return dt.astimezone(local_tz).strftime('%H:%M')
            return '--:--'

        # 4. Fill Data
        for row_num, emp in enumerate(employees, 2):
            ws.cell(row=row_num, column=1, value=emp.name)
            ws.cell(row=row_num, column=2, value=emp.employee_code)

            total_present, total_half, total_leave, total_absent, total_hrs = 0, 0, 0, 0, 0.0

            for i, current_date in enumerate(date_range):
                col_index = 3 + i
                record = attendance_dict.get((emp.id, current_date))
                
                if record:
                    if record.status == 'Present': total_present += 1
                    elif record.status == 'Half Day': total_half += 1
                    elif record.status == 'Leave': total_leave += 1
                    elif record.status in ['Absent', 'Mismatch']: total_absent += 1

                    if record.total_hours:
                        total_hrs += float(record.total_hours)

                    if record.status in ['Leave', 'Absent', 'Mismatch']:
                        cell_val = record.status
                    else:
                        in_t = fmt_time(record.sign_in_time)
                        out_t = fmt_time(record.sign_out_time)
                        hts = float(record.total_hours) if record.total_hours else 0
                        cell_val = f"{in_t} TO {out_t}\n({hts:.1f}h)"
                else:
                    total_absent += 1
                    cell_val = 'Absent' if current_date.weekday() < 5 else 'Weekend'

                cell = ws.cell(row=row_num, column=col_index, value=cell_val)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            summary_start = len(date_range) + 3
            ws.cell(row=row_num, column=summary_start, value=total_present).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=summary_start+1, value=total_half).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=summary_start+2, value=total_leave).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=summary_start+3, value=total_absent).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=summary_start+4, value=f"{total_hrs:.1f}h").alignment = Alignment(horizontal='center')

        # Generate response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        safe_filename = f"attendance_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{safe_filename}"'
        wb.save(response)
        return response
